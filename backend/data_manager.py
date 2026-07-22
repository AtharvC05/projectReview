# simplified_data_manager.py

from flask import Blueprint, render_template, request, jsonify, send_file, session, redirect, url_for
import logging
import io
import pandas as pd
import mysql.connector
from datetime import datetime
import re
import json
import base64
import openpyxl
import os
import hashlib
import backend.db as db
import backend.auth as auth

admin_required = auth.admin_required

logger = logging.getLogger(__name__)


bp = Blueprint('data_manager', __name__, template_folder='templates')

@bp.before_request
def require_admin():
    """Ensure all data manager routes require admin access"""
    if 'user_id' not in session:
        # For HTML routes, redirect to login
        if request.path.endswith('/data-manager'):
            return redirect(url_for('auth.login'))
        # For API routes, return JSON error
        return jsonify({'error': 'Authentication required'}), 401
    
    if session.get('role') != 'admin':
        # For HTML routes and API routes, return error
        if request.path.endswith('/data-manager'):
            return jsonify({'error': 'Admin access required'}), 403
        # For API routes, return JSON error
        return jsonify({'error': 'Admin access required'}), 403

# Single file storage for admin (no user sessions)
UPLOAD_FOLDER = 'admin_files'

def get_admin_file_path():
    year = db.get_academic_year()
    return os.path.join(UPLOAD_FOLDER, f'project_data_{year}.xlsx')

def get_admin_metadata_path():
    year = db.get_academic_year()
    return os.path.join(UPLOAD_FOLDER, f'metadata_{year}.json')

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = name.lower()
    name = name.replace("\n", " ").replace("\r", " ")
    name = name.replace(".", "")
    name = re.sub(r"\b(dr|prof|professor)\b", "", name)
    name = re.sub(r"\b[a-z]\b", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()

def normalize_column_name(column_name: str) -> str:
    """Normalize column names to handle various naming conventions"""
    if not column_name or pd.isna(column_name):
        return None
    
    col_name = str(column_name).strip().lower()
    col_name = re.sub(r'[_\-\s]+', ' ', col_name)
    col_name = re.sub(r'[^\w\s]', '', col_name)
    col_name = re.sub(r'\s+', ' ', col_name).strip()
    
    # Group No. variations
    if any(pattern in col_name for pattern in [
        'group no', 'grp no', 'group number', 'group id', 'grp id'
    ]):
        return 'Group No.'
    
    # Roll No. variations
    if any(pattern in col_name for pattern in [
        'roll no', 'roll number', 'rollno', 'roll', 'student id'
    ]):
        return 'Roll No.'
    
    # Student Name variations
    if any(pattern in col_name for pattern in [
        'name of the group member', 'name of group member', 'group member name',
        'name of members', 'member name', 'student name', 'name'
    ]) and 'guide' not in col_name and 'company' not in col_name and 'panel' not in col_name:
        return 'Name of the group member'
    
    # Contact Details variations
    if any(pattern in col_name for pattern in [
        'contact details', 'contact', 'phone', 'mobile', 'email', 'contact info'
    ]):
        return 'Contact details'
    
    # Project Domain variations
    if any(pattern in col_name for pattern in [
        'project domain', 'projects domain', 'domain', 'field', 'area'
    ]):
        return 'Project Domain'
    
    # Project Title variations
    if any(pattern in col_name for pattern in [
        'title of the project', 'project title', 'title', 'final title',
        'project name', 'title of project'
    ]):
        return 'Title of the Project'
    
    # Sponsor Company variations
    if any(pattern in col_name for pattern in [
        'name of the sponsored company', 'sponsored company', 'sponsor company',
        'company name', 'sponsoring company', 'name of sponsored'
    ]):
        return 'Name of the sponsored company '
    
    # Guide Name variations
    if any(pattern in col_name for pattern in [
        'name of the guide', 'guide name', 'guide', 'supervisor',
        'mentor name', 'faculty guide'
    ]):
        return 'Name of the Guide'
    
    # Schedule-related columns
    if any(pattern in col_name for pattern in ['track', 'panel no', 'panel number']):
        return 'Track'
    
    if any(pattern in col_name for pattern in [
        'name of the panel', 'panel name', 'panel members', 'panel professors',
        'evaluators', 'panel faculty'
    ]):
        return 'Name of the Panel'
    
    if any(pattern in col_name for pattern in [
        'group id', 'groups', 'group nos', 'assigned groups'
    ]):
        return 'Group ID'
    
    if any(pattern in col_name for pattern in [
        'location', 'room', 'venue', 'place'
    ]):
        return 'Location'
    
    return column_name

def load_excel_sheet_with_smart_header(xls, sheet_name, is_schedule=False):
    """Dynamically locate header row in Excel sheet and parse DataFrame"""
    try:
        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        header_row_idx = 0
        target_keywords = ['TRACK', 'PANEL', 'GROUP ID', 'LOCATION'] if is_schedule else ['GROUP NO', 'ROLL NO', 'MEMBER', 'DOMAIN', 'TITLE', 'GUIDE']
        
        for idx, row in df_raw.iloc[:10].iterrows():
            row_str = " ".join([str(val).upper() for val in row.values if pd.notnull(val)])
            if any(kw in row_str for kw in target_keywords):
                header_row_idx = idx
                break
                
        df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=header_row_idx)
        return df
    except Exception as e:
        logger.warning(f"Smart header detection fallback for {sheet_name}: {e}")
        fallback_skip = 2 if is_schedule else 3
        return pd.read_excel(xls, sheet_name=sheet_name, skiprows=fallback_skip)

def normalize_sheet_name(sheet_name: str) -> str:
    """Normalize sheet names to identify division and schedule sheets"""
    if not sheet_name:
        return sheet_name
    
    sheet_upper = sheet_name.upper().strip()
    
    # Division A variations
    if any(keyword in sheet_upper for keyword in [
        'FINAL DIV A', 'FINAL DIV-A', 'FINALDIVA', 'FINAL_DIV_A',
        'DIV A', 'DIV-A', 'DIVA', 'DIVISION A', 'DIVISION-A'
    ]):
        return 'div_a'
    
    # Division B variations
    if any(keyword in sheet_upper for keyword in [
        'FINAL DIV B', 'FINAL DIV-B', 'FINALDIVB', 'FINAL_DIV_B',
        'DIV B', 'DIV-B', 'DIVB', 'DIVISION B', 'DIVISION-B'
    ]):
        return 'div_b'
    
    # Schedule variations
    if any(keyword in sheet_upper for keyword in [
        'SCHEDULE', 'SCHED', 'PANEL SCHEDULE', 'EVALUATION SCHEDULE'
    ]):
        return 'schedule'
    
    return sheet_name

def normalize_dataframe_columns(df):
    """Normalize all column names in a DataFrame"""
    if df is None or df.empty:
        return df
    
    column_mapping = {}
    for col in df.columns:
        normalized = normalize_column_name(col)
        if normalized:
            column_mapping[col] = normalized
    
    df = df.rename(columns=column_mapping)
    return df

def detect_and_normalize_sheets_robust(xls):
    """Robust sheet detection that handles edge cases"""
    sheet_names = xls.sheet_names
    detected_sheets = {
        'div_a': None,
        'div_b': None, 
        'schedule': None
    }
    
    logger.info(f"Available sheets: {sheet_names}")
    
    for sheet_name in sheet_names:
        sheet_upper = sheet_name.upper().strip()
        logger.info(f"Processing sheet: '{sheet_name}' -> '{sheet_upper}'")
        
        # Check for Division A
        if any(pattern in sheet_upper for pattern in [
            'FINAL DIV A', 'FINALDIVA', 'FINAL_DIV_A', 'FINAL-DIV-A',
            'DIV A', 'DIVA', 'DIV-A', 'DIV_A', 
            'DIVISION A', 'DIVISIONA', 'DIVISION-A', 'DIVISION_A'
        ]):
            detected_sheets['div_a'] = sheet_name
            logger.info(f"Detected Division A sheet: {sheet_name}")
            
        # Check for Division B  
        elif any(pattern in sheet_upper for pattern in [
            'FINAL DIV B', 'FINALDIVB', 'FINAL_DIV_B', 'FINAL-DIV-B',
            'DIV B', 'DIVB', 'DIV-B', 'DIV_B',
            'DIVISION B', 'DIVISIONB', 'DIVISION-B', 'DIVISION_B'
        ]):
            detected_sheets['div_b'] = sheet_name
            logger.info(f"Detected Division B sheet: {sheet_name}")
            
        # Check for Schedule
        elif any(pattern in sheet_upper for pattern in [
            'SCHEDULE', 'SCHED', 'PANEL SCHEDULE', 'EVALUATION SCHEDULE',
            'PANEL_SCHEDULE', 'EVALUATION_SCHEDULE', 'PANEL-SCHEDULE'
        ]):
            detected_sheets['schedule'] = sheet_name
            logger.info(f"Detected Schedule sheet: {sheet_name}")
    
    # Fallback logic for unassigned sheets if exact keywords were not matched
    used_sheets = set(v for v in detected_sheets.values() if v is not None)
    remaining_sheets = [s for s in sheet_names if s not in used_sheets]
    
    if detected_sheets['div_a'] is None and remaining_sheets:
        detected_sheets['div_a'] = remaining_sheets.pop(0)
    if detected_sheets['div_b'] is None and remaining_sheets:
        detected_sheets['div_b'] = remaining_sheets.pop(0)
    if detected_sheets['schedule'] is None and remaining_sheets:
        detected_sheets['schedule'] = remaining_sheets.pop(0)
    elif detected_sheets['schedule'] is None and sheet_names:
        detected_sheets['schedule'] = sheet_names[-1]

    logger.info(f"Final detection results: {detected_sheets}")
    return detected_sheets

def save_admin_file(file_content, metadata):
    """Save Excel file and metadata for admin"""
    try:
        # Save Excel file
        with open(get_admin_file_path(), 'wb') as f:
            f.write(file_content)
        
        # Save metadata
        with open(get_admin_metadata_path(), 'w') as f:
            json.dump(metadata, f)
        
        return True
    except Exception as e:
        logger.error(f"Error saving admin file: {e}")
        return False

def load_admin_file():
    """Load Excel file and metadata for admin"""
    try:
        file_path = get_admin_file_path()
        metadata_path = get_admin_metadata_path()
        if not os.path.exists(file_path) or not os.path.exists(metadata_path):
            return None, None
        
        # Load Excel file
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        # Load metadata
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)
        
        return file_content, metadata
    except Exception as e:
        logger.error(f"Error loading admin file: {e}")
        return None, None


def extract_all_group_ids(row_data):
    """Extract group IDs from row data"""
    all_groups = set()
    
    for cell_value in row_data:
        if not cell_value:
            continue
        
        cell_str = str(cell_value).upper().strip()
        matches = re.findall(r'\b(BI[AB][-_\s]*\d{1,2})\b', cell_str)
        for match in matches:
            normalized = db.normalize_group_id(match)
            if normalized:
                all_groups.add(normalized)
    
    return sorted(list(all_groups))

def parse_panel_professors(panel_text: str):
    """Parse panel professor names handling multiline, merged rows, slashes, commas, and numbered lists"""
    if not panel_text or pd.isnull(panel_text) or str(panel_text).strip().lower() == 'nan':
        return []
    text = str(panel_text).replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'[/;,|&]', '\n', text)
    lines = text.split('\n')
    panel_profs = []
    for line in lines:
        clean = line.strip()
        clean = re.sub(r'^[\d\w][.\)]\s*', '', clean).strip()
        clean = re.sub(r'^[•\-\*]\s*', '', clean).strip()
        if len(clean) >= 3 and any(c.isalpha() for c in clean) and not clean.lower().startswith('track') and not clean.lower().startswith('lab'):
            panel_profs.append(clean)
    return panel_profs

def assign_evaluators_from_panel(panel_professors, group_ids):
    """Assign evaluators from panel professors"""
    if not panel_professors or len(panel_professors) < 2:
        panel_professors = ["Default Prof 1", "Default Prof 2", "Default Prof 3"]
    
    assignments = []
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    
    for i, group_id in enumerate(group_ids):
        db_group_id = db.add_year_prefix(group_id)
        cur.execute("SELECT guide_name FROM projects WHERE group_id = %s", (db_group_id,))
        proj_row = cur.fetchone()
        guide_name = proj_row['guide_name'] if proj_row else ""

        available_evaluators = [prof for prof in panel_professors 
                              if normalize_name(prof) != normalize_name(guide_name)]
        
        eval1 = available_evaluators[0] if len(available_evaluators) > 0 else "Default Eval 1"
        eval2 = available_evaluators[1] if len(available_evaluators) > 1 else "Default Eval 2"
        
        assignments.append({
            'group_id': db_group_id,
            'guide': guide_name or "",
            'evaluator1': eval1 or "",
            'evaluator2': eval2 or ""
        })
    
    cur.close()
    conn.close()
    return assignments

def process_division_enhanced_with_normalization(df, division_name):
    """Enhanced division processing with column normalization"""
    if df is None or df.empty:
        return 0, 0
    
    # Normalize column names first
    df = normalize_dataframe_columns(df)
    
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    
    group_id = None
    processed_groups = []
    processed_members = 0
    
    for i, row in df.iterrows():
        try:
            # Use normalized column names - handle Series
            group_no_value = row.get('Group No.', '')
            # Convert Series to scalar if needed
            if isinstance(group_no_value, pd.Series):
                group_no_value = group_no_value.iloc[0] if not group_no_value.empty else ''
            
            if pd.notnull(group_no_value) and str(group_no_value).strip():
                group_id = str(group_no_value).strip()
                
                # Normalize group ID format so BIA01, BIA1, BIA-1 all become BIA-01
                group_id = db.normalize_group_id(group_id)
                
                # Insert project data with normalized column access
                try:
                    # Handle potential Series for all project fields
                    project_domain_val = row.get('Project Domain', '')
                    if isinstance(project_domain_val, pd.Series):
                        project_domain_val = project_domain_val.iloc[0] if not project_domain_val.empty else ''
                    project_domain = str(project_domain_val).strip()[:255] if pd.notnull(project_domain_val) else ""
                    
                    project_title_val = row.get('Title of the Project', '')
                    if isinstance(project_title_val, pd.Series):
                        project_title_val = project_title_val.iloc[0] if not project_title_val.empty else ''
                    project_title = str(project_title_val).strip()[:500] if pd.notnull(project_title_val) else ""
                    
                    sponsor_company_val = row.get('Name of the sponsored company ', '')
                    if isinstance(sponsor_company_val, pd.Series):
                        sponsor_company_val = sponsor_company_val.iloc[0] if not sponsor_company_val.empty else ''
                    sponsor_company = str(sponsor_company_val).strip()[:255] if pd.notnull(sponsor_company_val) else ""
                    
                    guide_name_val = row.get('Name of the Guide', '')
                    if isinstance(guide_name_val, pd.Series):
                        guide_name_val = guide_name_val.iloc[0] if not guide_name_val.empty else ''
                    guide_name = str(guide_name_val).strip()[:100] if pd.notnull(guide_name_val) else ""
                    
                    db_group_id = db.add_year_prefix(group_id)
                    cur.execute(
                        """INSERT INTO projects 
                           (group_id, division, project_domain, project_title, sponsor_company, guide_name, 
                            evaluator1_name, evaluator2_name) 
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE
                           division = VALUES(division),
                           project_domain = VALUES(project_domain),
                           project_title = VALUES(project_title),
                           sponsor_company = VALUES(sponsor_company),
                           guide_name = VALUES(guide_name)""",
                        (db_group_id, division_name, project_domain, project_title, sponsor_company, guide_name, "", "")
                    )
                    
                    if group_id not in processed_groups:
                        processed_groups.append(group_id)
                        
                except Exception as project_error:
                    print(f"Error inserting project {group_id}: {project_error}", flush=True)
                    logger.error(f"Error inserting project {group_id}: {str(project_error)}")
            
            # Process member data with normalized column access
            # Handle potential Series by using .iloc[0] to get scalar values
            roll_no_val = row.get('Roll No.', '')
            student_name_val = row.get('Name of the group member', '')
            
            # Convert Series to scalar if needed
            if isinstance(roll_no_val, pd.Series):
                roll_no_val = roll_no_val.iloc[0] if not roll_no_val.empty else ''
            if isinstance(student_name_val, pd.Series):
                student_name_val = student_name_val.iloc[0] if not student_name_val.empty else ''
            
            if (group_id and 
                pd.notnull(roll_no_val) and str(roll_no_val).strip() and
                pd.notnull(student_name_val) and str(student_name_val).strip()):
                
                try:
                    roll_no = str(roll_no_val).strip()
                    student_name = str(student_name_val).strip()[:100]
                    
                    if roll_no and student_name:
                        db_group_id = db.add_year_prefix(group_id)
                        db_roll_no = db.add_year_prefix(roll_no)
                        cur.execute(
                            """INSERT INTO members (group_id, roll_no, student_name) 
                               VALUES (%s, %s, %s)
                               ON DUPLICATE KEY UPDATE
                               student_name = VALUES(student_name),
                               group_id = VALUES(group_id)""",
                            (db_group_id, db_roll_no, student_name)
                        )
                        processed_members += 1

                        
                except Exception as member_error:
                    print(f"Error inserting member for {group_id}: {member_error}", flush=True)
                    logger.error(f"Error inserting member for {group_id}: {str(member_error)}")
                
        except Exception as row_error:
            print(f"Error processing row {i} in {division_name}: {row_error}", flush=True)
            logger.warning(f"Error processing row {i} in {division_name}: {str(row_error)}")
            continue
            
    conn.commit()
    cur.close()
    conn.close()
    return len(processed_groups), processed_members

def process_all_data_with_normalization(div_a, div_b, sched):
    """Process all data with column normalization"""
    conn = db.get_connection()
    cur = conn.cursor(dictionary=True)
    
    # Process divisions with normalization without deleting existing records
    # to preserve attendance, generated PDFs, review marks, and final summary comments
    active_year = db.get_academic_year()
    
    # Process divisions with normalization
    div_a_groups, div_a_members = process_division_enhanced_with_normalization(div_a, 'A')
    div_b_groups, div_b_members = process_division_enhanced_with_normalization(div_b, 'B')
    
    # Process schedule with Track block grouping (handling merged Excel rows)
    schedule_processed = 0
    all_scheduled_groups = set()
    
    # Normalize schedule DataFrame
    sched = normalize_dataframe_columns(sched)
    
    # Step 1: Group rows by Track block
    track_blocks = []
    current_track = None
    panel_texts_acc = []
    row_values_acc = []
    location_acc = ""

    for i, row in sched.iterrows():
        track_val = row.get('Track')
        # Check if row starts a new track block (e.g. 1, 2, 3...)
        if pd.notnull(track_val) and str(track_val).strip().replace('.', '').isdigit():
            if current_track is not None:
                track_blocks.append({
                    'track': current_track,
                    'panel_texts': panel_texts_acc,
                    'row_values': row_values_acc,
                    'location': location_acc
                })
            current_track = int(float(str(track_val).strip()))
            panel_texts_acc = []
            row_values_acc = []
            location_acc = ""
            
        if current_track is not None:
            p_text = str(row.get('Name of the Panel', ''))
            if p_text and p_text.lower() != 'nan':
                panel_texts_acc.append(p_text)
                
            loc = str(row.get('Location', '')).strip()
            if loc and loc.lower() != 'nan':
                location_acc = loc
                
            row_values = [str(val) for val in row.values if pd.notnull(val)]
            row_values_acc.extend(row_values)

    if current_track is not None:
        track_blocks.append({
            'track': current_track,
            'panel_texts': panel_texts_acc,
            'row_values': row_values_acc,
            'location': location_acc
        })

    # Step 2: Process each Track block
    for block in track_blocks:
        try:
            track = block['track']
            panel_profs = parse_panel_professors("\n".join(block['panel_texts']))
            if not panel_profs:
                panel_profs = [f"Default Panel {track} Prof 1", f"Default Panel {track} Prof 2", f"Default Panel {track} Prof 3"]
                
            location = block['location'] if block['location'] else f"Room {track}"
            group_ids = extract_all_group_ids(block['row_values'])
            
            if not group_ids:
                continue
                
            assignments = assign_evaluators_from_panel(panel_profs, group_ids)
            
            for assignment in assignments:
                try:
                    group_id = assignment['group_id']
                    cur.execute("""
                        INSERT INTO panel_assignments
                        (group_id, track, panel_professors, location, guide, reviewer1, reviewer2, reviewer3)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        track=VALUES(track), panel_professors=VALUES(panel_professors), 
                        location=VALUES(location), guide=VALUES(guide),
                        reviewer1=VALUES(reviewer1), reviewer2=VALUES(reviewer2)
                    """, (
                        group_id, track, '\n'.join(panel_profs), location,
                        assignment['guide'], assignment['evaluator1'], assignment['evaluator2'], None
                    ))
                    
                    cur.execute("""
                        UPDATE projects 
                        SET evaluator1_name = %s, evaluator2_name = %s 
                        WHERE group_id = %s
                    """, (assignment['evaluator1'], assignment['evaluator2'], group_id))
                    
                    all_scheduled_groups.add(group_id)
                    
                except Exception as assignment_error:
                    print(f"Error assigning {assignment['group_id']}: {assignment_error}", flush=True)
                    logger.error(f"Error assigning {assignment['group_id']}: {str(assignment_error)}")
            
            schedule_processed += 1
            
        except Exception as e:
            print(f"Error processing schedule block for track {block.get('track')}: {e}", flush=True)
            logger.warning(f"Error processing schedule block for track {block.get('track')}: {str(e)}")
            continue
    
    conn.commit()
    cur.close()
    conn.close()
    
    return {
        'div_a_groups': div_a_groups,
        'div_b_groups': div_b_groups,
        'scheduled_groups': len(all_scheduled_groups),
        'total_processed': div_a_groups + div_b_groups
    }

def generate_cell_mapping(div_a, div_b, sched, div_a_sheet, div_b_sheet, schedule_sheet):
    """Generate mapping between visual cells and database fields"""
    mapping = {
        'div_a': {},
        'div_b': {},
        'schedule': {}
    }
    
    # Map Division A editable fields
    for i, row in div_a.iterrows():
        actual_row = i + 4  # Account for skipped rows
        if pd.notnull(row.get('Roll No.', '')):
            mapping['div_a'][f'{actual_row}'] = {
                'roll_no': {'col': 'B', 'value': str(row.get('Roll No.', ''))},
                'student_name': {'col': 'C', 'value': str(row.get('Name of the group member', ''))},
            }
    
    # Map Division B editable fields  
    for i, row in div_b.iterrows():
        actual_row = i + 4
        if pd.notnull(row.get('Roll No.', '')):
            mapping['div_b'][f'{actual_row}'] = {
                'roll_no': {'col': 'B', 'value': str(row.get('Roll No.', ''))},
                'student_name': {'col': 'C', 'value': str(row.get('Name of the group member', ''))},
            }
    
    # Map Schedule editable fields
    for i, row in sched.iterrows():
        actual_row = i + 3
        if pd.notnull(row.get('Track', '')):
            mapping['schedule'][f'{actual_row}'] = {
                'track': {'col': 'A', 'value': str(row.get('Track', ''))},
                'panel_professors': {'col': 'B', 'value': str(row.get('Name of the Panel', ''))},
                'group_ids': {'col': 'C', 'value': str(row.get('Group ID', ''))},
                'location': {'col': 'D', 'value': str(row.get('Location', ''))}
            }
    
    return mapping

@bp.route('/data-manager')

def data_manager_page():
    return render_template('data-manager.html')


@bp.route('/api/check-stored-file', methods=['GET'])
def check_stored_file():
    """Check if admin has a stored file"""
    try:
        file_content, metadata = load_admin_file()
        
        if file_content and metadata:
            return jsonify({
                'has_stored_file': True,
                'metadata': metadata,
                'upload_date': metadata.get('upload_date', 'Unknown')
            })
        else:
            return jsonify({'has_stored_file': False})
            
    except Exception as e:
        logger.error(f"Error checking stored file: {e}")
        return jsonify({'has_stored_file': False})

@bp.route('/api/load-stored-file', methods=['POST'])
def load_stored_file():
    """Load admin's stored file"""
    try:
        file_content, metadata = load_admin_file()
        
        if not file_content or not metadata:
            return jsonify({'success': False, 'error': 'No stored file found'}), 404
        
        # Convert file content to base64
        original_b64 = base64.b64encode(file_content).decode('utf-8')
        
        # Process the stored file
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        # Get sheet names from metadata
        sheet_names = metadata.get('sheet_names', {})
        
        if not all([sheet_names.get('div_a'), sheet_names.get('div_b'), sheet_names.get('schedule')]):
            return jsonify({'success': False, 'error': 'Invalid stored file format'}), 400

        # Load and process data with normalization
        div_a = load_excel_sheet_with_smart_header(xls, sheet_names['div_a'], is_schedule=False)
        div_b = load_excel_sheet_with_smart_header(xls, sheet_names['div_b'], is_schedule=False)
        sched = load_excel_sheet_with_smart_header(xls, sheet_names['schedule'], is_schedule=True)

        # Normalize column names
        div_a = normalize_dataframe_columns(div_a)
        div_b = normalize_dataframe_columns(div_b)
        sched = normalize_dataframe_columns(sched)

        # Process with normalization
        extracted_data = process_all_data_with_normalization(div_a, div_b, sched)
        
        # Generate cell mapping for editing
        cell_mapping = generate_cell_mapping(div_a, div_b, sched, 
                                           sheet_names['div_a'], 
                                           sheet_names['div_b'], 
                                           sheet_names['schedule'])
        
        return jsonify({
            'success': True,
            'original_file': original_b64,
            'extracted_data': extracted_data,
            'cell_mapping': cell_mapping,
            'sheet_names': sheet_names,
            'message': f'Loaded stored file from {metadata.get("upload_date", "Unknown")}'
        })
        
    except Exception as e:
        logger.error(f"Error loading stored file: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/import-excel', methods=['POST'])
def import_excel():
    """Import Excel and return both visual representation and extracted data"""
    try:
        file = request.files.get('excel')
        if not file:
            return jsonify({'success': False, 'error': 'No file provided'}), 400

        file_content = file.read()
        
        # Store original file for visual rendering
        original_b64 = base64.b64encode(file_content).decode('utf-8')
        
        # Extract data using enhanced normalization
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        # Use the more robust sheet detection
        detected_sheets = detect_and_normalize_sheets_robust(xls)
        
        # Check if all required sheets were found
        missing_sheets = [key for key, value in detected_sheets.items() if value is None]
        if missing_sheets:
            available_sheets = list(xls.sheet_names)
            return jsonify({
                'success': False,
                'error': f"Could not detect {missing_sheets} sheets. Available sheets: {available_sheets}. Detected: {detected_sheets}"
            }), 400

        # Prepare metadata
        metadata = {
            'sheet_names': {
                'div_a': detected_sheets['div_a'],
                'div_b': detected_sheets['div_b'],
                'schedule': detected_sheets['schedule']
            },
            'upload_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'original_filename': file.filename
        }
        
        # Save file for admin
        save_success = save_admin_file(file_content, metadata)
        
        # Load and normalize data
        div_a = load_excel_sheet_with_smart_header(xls, detected_sheets['div_a'], is_schedule=False)
        div_b = load_excel_sheet_with_smart_header(xls, detected_sheets['div_b'], is_schedule=False)
        sched = load_excel_sheet_with_smart_header(xls, detected_sheets['schedule'], is_schedule=True)
        
        # Normalize column names
        div_a = normalize_dataframe_columns(div_a)
        div_b = normalize_dataframe_columns(div_b)
        sched = normalize_dataframe_columns(sched)
        
        # Process with enhanced normalization
        extracted_data = process_all_data_with_normalization(div_a, div_b, sched)
        
        # Generate cell mapping
        cell_mapping = generate_cell_mapping(div_a, div_b, sched, 
                                           detected_sheets['div_a'], 
                                           detected_sheets['div_b'], 
                                           detected_sheets['schedule'])
        
        message = 'Excel imported successfully with column normalization and visual preservation'
        if save_success:
            message += ' and saved for future access'
        
        return jsonify({
            'success': True,
            'original_file': original_b64,
            'extracted_data': extracted_data,
            'cell_mapping': cell_mapping,
            'sheet_names': metadata['sheet_names'],
            'message': message,
            'normalized_columns': {
                'div_a': list(div_a.columns),
                'div_b': list(div_b.columns),
                'schedule': list(sched.columns)
            }
        })
        
    except Exception as e:
        print(f"❌ Import error in data_manager: {e}", flush=True)
        logger.error(f"Import error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/update-cell-general', methods=['POST'])
def update_cell_general():
    """Update any cell in the Excel sheets and sync with database"""
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        row = data.get('row')
        col = data.get('col')
        new_value = data.get('value', '')
        old_value = data.get('old_value', '')
        
        # Store the update in database
        conn = db.get_connection()
        cur = conn.cursor()
        
        # Create a general cell updates table if it doesn't exist
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cell_updates (
                id INT AUTO_INCREMENT PRIMARY KEY,
                sheet_name VARCHAR(100),
                row_num INT,
                col_num INT,
                old_value TEXT,
                new_value TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_cell (sheet_name, row_num, col_num)
            )
        """)
        
        # Store the cell update
        cur.execute("""
            INSERT INTO cell_updates (sheet_name, row_num, col_num, old_value, new_value)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
            old_value = VALUES(old_value),
            new_value = VALUES(new_value),
            updated_at = CURRENT_TIMESTAMP
        """, (sheet_name, row, col, old_value, new_value))
        
        # Try to update specific database tables based on sheet and content
        try:
            update_specific_database_field(cur, sheet_name, row, col, new_value)
        except Exception as e:
            logger.warning(f"Could not update specific database field: {e}")
        
        conn.commit()
        
        # Update stored admin file
        update_admin_file_cell(sheet_name, row, col, new_value)
        
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Updated cell [{row},{col}] in {sheet_name}'
        })
        
    except Exception as e:
        logger.error(f"General cell update error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

def update_admin_file_cell(sheet_name, row, col, new_value):
    """Update a cell in the stored admin Excel file"""
    try:
        file_path = get_admin_file_path()
        if not os.path.exists(file_path):
            return
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Find the correct worksheet
        ws = None
        for ws_name in wb.sheetnames:
            if ws_name.upper() == sheet_name.upper():
                ws = wb[ws_name]
                break
        
        if ws:
            # Update the cell (note: openpyxl uses 1-based indexing)
            ws.cell(row=row + 1, column=col + 1, value=new_value)
            
            # Save the workbook
            wb.save(file_path)
            
    except Exception as e:
        logger.warning(f"Could not update admin file: {e}")

def update_specific_database_field(cursor, sheet_name, row, col, new_value):
    """Update specific database fields based on sheet position"""
    sheet_upper = sheet_name.upper()
    
    # Handle Division A/B sheets
    if any(keyword in sheet_upper for keyword in ['DIV A', 'DIVA', 'DIVISION A']):
        division = 'A'
        update_division_field(cursor, division, row, col, new_value)
    elif any(keyword in sheet_upper for keyword in ['DIV B', 'DIVB', 'DIVISION B']):
        division = 'B'
        update_division_field(cursor, division, row, col, new_value)
    elif any(keyword in sheet_upper for keyword in ['SCHEDULE', 'SCHED']):
        update_schedule_field(cursor, row, col, new_value)

def update_division_field(cursor, division, row, col, new_value):
    """Update fields in projects/members tables based on position"""
    # Common column mappings for division sheets
    if row < 1:  # Skip header rows
        return
        
    # Approximate column mappings
    column_mappings = {
        0: 'group_id',      # Group No.
        1: 'roll_no',       # Roll No.
        2: 'student_name',  # Name of group member
        3: 'project_domain', # Project Domain
        4: 'project_title',  # Title of Project
        5: 'sponsor_company', # Sponsor Company
        6: 'guide_name'      # Guide Name
    }
    
    field_name = column_mappings.get(col)
    if not field_name:
        return
    
    try:
        active_year = db.get_academic_year()
        if field_name == 'group_id':
            # Update projects table directly
            cursor.execute("""
                UPDATE projects SET group_id = %s 
                WHERE division = %s AND group_id = (
                    SELECT old_group_id FROM (
                        SELECT group_id as old_group_id FROM projects 
                        WHERE division = %s AND group_id LIKE %s LIMIT 1 OFFSET %s
                    ) as temp
                )
            """, (db.add_year_prefix(new_value), division, division, f"{active_year}_%", row - 1))
        elif field_name == 'roll_no':
            # Update members table with prefixed roll number
            cursor.execute("""
                UPDATE members m
                JOIN projects p ON m.group_id = p.group_id
                SET m.roll_no = %s
                WHERE p.division = %s AND p.group_id LIKE %s
                LIMIT 1 OFFSET %s
            """, (db.add_year_prefix(new_value), division, f"{active_year}_%", row - 1))
        elif field_name == 'student_name':
            # Update members table
            cursor.execute("""
                UPDATE members m
                JOIN projects p ON m.group_id = p.group_id
                SET m.student_name = %s
                WHERE p.division = %s AND p.group_id LIKE %s
                LIMIT 1 OFFSET %s
            """, (new_value, division, f"{active_year}_%", row - 1))
        elif field_name in ['project_domain', 'project_title', 'sponsor_company', 'guide_name']:
            # Update projects table
            cursor.execute("""
                UPDATE projects 
                SET {} = %s 
                WHERE division = %s AND group_id LIKE %s
                LIMIT 1 OFFSET %s
            """.format(field_name), (new_value, division, f"{active_year}_%", row - 1))
            
    except Exception as e:
        logger.warning(f"Could not update {field_name} in division {division}: {e}")

def update_schedule_field(cursor, row, col, new_value):
    """Update schedule/panel assignments based on position"""
    if row < 1:  # Skip headers
        return
        
    # Schedule column mappings
    schedule_mappings = {
        0: 'track',
        1: 'panel_professors',
        2: 'group_ids',
        3: 'location'
    }
    
    field_name = schedule_mappings.get(col)
    if not field_name:
        return
        
    try:
        active_year = db.get_academic_year()
        if field_name == 'track':
            cursor.execute("""
                UPDATE panel_assignments 
                SET track = %s 
                WHERE group_id LIKE %s AND track = (
                    SELECT old_track FROM (
                        SELECT track as old_track FROM panel_assignments 
                        WHERE group_id LIKE %s
                        ORDER BY track LIMIT 1 OFFSET %s
                    ) as temp
                )
            """, (new_value, f"{active_year}_%", f"{active_year}_%", row - 1))
        elif field_name in ['panel_professors', 'location']:
            cursor.execute("""
                UPDATE panel_assignments 
                SET {} = %s 
                WHERE group_id LIKE %s
                ORDER BY track LIMIT 1 OFFSET %s
            """.format(field_name), (new_value, f"{active_year}_%", row - 1))
            
    except Exception as e:
        logger.warning(f"Could not update {field_name} in schedule: {e}")

@bp.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export current admin file as Excel"""
    try:
        file_path = get_admin_file_path()
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'No file to export'}), 400
        
        # Return the updated admin file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=f'updated_project_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/export-formatted-excel', methods=['POST'])
def export_formatted_excel():
    """Export Excel file with original formatting and merged cells preserved"""
    try:
        file_path = get_admin_file_path()
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'No stored file found'}), 404
        
        # Load the original workbook with all formatting preserved
        wb = openpyxl.load_workbook(file_path)
        
        # Get database data for updates
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT p.group_id, p.division, p.project_domain, p.project_title, 
                   p.sponsor_company, p.guide_name, p.evaluator1_name, p.evaluator2_name,
                   COALESCE(m.roll_no, '') as roll_no, 
                   COALESCE(m.student_name, '') as student_name
            FROM projects p 
            LEFT JOIN members m ON p.group_id = m.group_id
            WHERE (p.group_id IS NOT NULL OR m.roll_no IS NOT NULL)
            ORDER BY p.division, p.group_id, m.roll_no
        """)
        database_data = cursor.fetchall()
        
        # Get panel assignments
        cursor.execute("""
            SELECT track, panel_professors, location, group_id,
                   guide, reviewer1, reviewer2
            FROM panel_assignments 
            ORDER BY track, group_id
        """)
        schedule_data = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Update each sheet while preserving formatting
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_upper = sheet_name.upper()
            
            if any(keyword in sheet_upper for keyword in ['DIV A', 'DIVA', 'DIVISION A']):
                update_division_sheet_formatted(ws, database_data, 'A')
            elif any(keyword in sheet_upper for keyword in ['DIV B', 'DIVB', 'DIVISION B']):
                update_division_sheet_formatted(ws, database_data, 'B')
            elif any(keyword in sheet_upper for keyword in ['SCHEDULE', 'SCHED']):
                update_schedule_sheet_formatted(ws, schedule_data)
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'formatted_project_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Formatted export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

def update_division_sheet_formatted(ws, data, division):
    """Update division sheet data while preserving all original formatting"""
    # Filter and group data by division
    division_data = [row for row in data if row['division'] == division]
    
    # Find the header row by looking for "Group No." or similar
    header_row = None
    data_start_row = 5  # Default fallback
    
    for row_num in range(1, min(15, ws.max_row + 1)):
        for col_num in range(1, 5):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value and any(keyword in str(cell_value).upper() for keyword in ['GROUP NO', 'ROLL NO', 'NAME OF']):
                header_row = row_num
                data_start_row = row_num + 1
                break
        if header_row:
            break
    
    current_row = data_start_row
    current_group = None
    group_start_row = None
    
    for row_data in division_data:
        # Only process rows that have student data
        if row_data.get('roll_no') and row_data.get('student_name'):
            try:
                # Check if we're starting a new group
                if row_data['group_id'] != current_group:
                    current_group = row_data['group_id']
                    group_start_row = current_row
                    
                    # Update group-level information
                    update_cell_preserve_format(ws, group_start_row, 1, row_data['group_id'])
                    update_cell_preserve_format(ws, group_start_row, 5, row_data.get('project_domain', ''))
                    update_cell_preserve_format(ws, group_start_row, 6, row_data.get('project_title', ''))
                    update_cell_preserve_format(ws, group_start_row, 7, row_data.get('sponsor_company', ''))
                    update_cell_preserve_format(ws, group_start_row, 8, row_data.get('guide_name', ''))
                
                # Update student-level information
                update_cell_preserve_format(ws, current_row, 2, row_data.get('roll_no', ''))
                update_cell_preserve_format(ws, current_row, 3, row_data.get('student_name', ''))
                
                current_row += 1
                
            except Exception as e:
                logger.warning(f"Error updating division {division} row {current_row}: {e}")
                current_row += 1
                continue

def update_schedule_sheet_formatted(ws, schedule_data):
    """Update schedule sheet while preserving formatting"""
    # Find the header row
    header_row = None
    data_start_row = 3  # Default
    
    for row_num in range(1, min(10, ws.max_row + 1)):
        for col_num in range(1, 5):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value and ('Track' in str(cell_value) or 'track' in str(cell_value).lower()):
                header_row = row_num
                data_start_row = row_num + 1
                break
        if header_row:
            break
    
    current_row = data_start_row
    
    # Group schedule data by track
    tracks = {}
    for schedule_row in schedule_data:
        track = schedule_row.get('track')
        if track is not None:
            if track not in tracks:
                tracks[track] = {
                    'track': track,
                    'panel_professors': schedule_row.get('panel_professors', ''),
                    'location': schedule_row.get('location', ''),
                    'group_ids': []
                }
            if schedule_row.get('group_id'):
                tracks[track]['group_ids'].append(schedule_row.get('group_id', ''))
    
    # Update each track
    for track_num in sorted(tracks.keys()):
        track_info = tracks[track_num]
        
        try:
            if current_row <= ws.max_row + 5:
                # Update track information
                update_cell_preserve_format(ws, current_row, 1, track_info['track'])
                
                # Clean up panel professors text
                panel_text = track_info['panel_professors']
                if panel_text:
                    panel_text = panel_text.replace('|', '\n').replace(',', '\n')
                    panel_lines = [line.strip() for line in panel_text.split('\n') if line.strip()]
                    panel_text = '\n'.join(panel_lines)
                
                update_cell_preserve_format(ws, current_row, 2, panel_text)
                
                # Join group IDs
                group_ids_str = ' '.join(track_info['group_ids']) if track_info['group_ids'] else ''
                update_cell_preserve_format(ws, current_row, 3, group_ids_str)
                
                update_cell_preserve_format(ws, current_row, 4, track_info['location'])
                
                current_row += 1
            else:
                logger.warning(f"Reached maximum rows, skipping track {track_num}")
                break
                
        except Exception as e:
            logger.warning(f"Error updating schedule track {track_num}: {e}")
            current_row += 1

def update_cell_preserve_format(ws, row, col, value):
    """Update cell value while preserving all original formatting"""
    try:
        if row <= 0 or col <= 0:
            return
            
        cell = ws.cell(row=row, column=col)
        
        # Store original formatting
        original_font = cell.font
        original_alignment = cell.alignment
        original_border = cell.border
        original_fill = cell.fill
        original_number_format = cell.number_format
        
        # Handle None values and clean up the value
        if value is None:
            value = ''
        elif not isinstance(value, (str, int, float)):
            value = str(value)
        
        # Clean up string values
        if isinstance(value, str):
            value = value.strip()
            if value.lower() in ['none', 'null', 'nan']:
                value = ''
        
        # Update value
        cell.value = value
        
        # Restore formatting
        from openpyxl.styles import Font, Alignment
        
        if original_font:
            cell.font = Font(
                name=original_font.name,
                size=original_font.size,
                bold=original_font.bold,
                italic=original_font.italic,
                vertAlign=original_font.vertAlign,
                underline=original_font.underline,
                strike=original_font.strike,
                color=original_font.color
            )
        
        if original_alignment:
            cell.alignment = Alignment(
                horizontal=original_alignment.horizontal,
                vertical=original_alignment.vertical,
                text_rotation=original_alignment.text_rotation,
                wrap_text=original_alignment.wrap_text,
                shrink_to_fit=original_alignment.shrink_to_fit,
                indent=original_alignment.indent
            )
        
        cell.border = original_border
        cell.fill = original_fill
        cell.number_format = original_number_format
        
    except Exception as e:
        logger.warning(f"Error updating cell {row},{col} with value '{value}': {e}")

@bp.route('/api/get-file-info', methods=['GET'])
def get_file_info():
    """Get information about stored admin file"""
    try:
        metadata_path = get_admin_metadata_path()
        if os.path.exists(metadata_path):
            with open(metadata_path, 'r') as f:
                metadata = json.load(f)
            return jsonify({'success': True, 'metadata': metadata})
        else:
            return jsonify({'success': False, 'error': 'No file information found'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})